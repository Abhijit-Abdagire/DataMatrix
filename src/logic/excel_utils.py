from openpyxl import Workbook, load_workbook
import pandas as pd
from src.logic.common_utils import logging, get_row_order_data
from src.logic.nse_utils import all_indices_data
import time


def write_indices_to_excel(indices_data, excel_path):

    try:
        logging.info("Saving data to Excel Sheet function triggered")
        ref_column_name = "B"
        row_order_map_data = get_row_order_data()

        if row_order_map_data:
            start_time = time.time()

            if ref_column_name in indices_data.columns:

                sorted_data = pd.concat(
                    [indices_data[indices_data[ref_column_name] == key] for key in row_order_map_data.keys()] +
                    [indices_data[~indices_data[ref_column_name].isin(row_order_map_data.keys())]]
                ).reset_index(drop=True)
            else:
                logging.warning(f"Column '{ref_column_name}' not found in DataFrame. Skipping rearrangement.")
                sorted_data = indices_data

            logging.info(f"Row Order Rearranging Time : {time.time() - start_time:.4f} seconds")

        else:
            sorted_data = indices_data

        excel_writing_start_time = time.time()
        excel_workbook = load_workbook(excel_path)

        if "AllIndicesData" in excel_workbook.sheetnames:
            sheet = excel_workbook["AllIndicesData"]
        else:
            sheet = excel_workbook.create_sheet("AllIndicesData")
            for col_idx, column in enumerate(sorted_data.columns, start=1):
                sheet.cell(row=1, column=col_idx, value=column)

        # Write the DataFrame to the sheet
        for row_idx, row in enumerate(sorted_data.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        excel_workbook.save(excel_path)
        logging.info(f"Data successfully saved to: {excel_path}")
        logging.info(f"Writing Indices data to Excel Time : {time.time() - excel_writing_start_time:.4f} seconds")
        # return True

    except Exception as e:
        logging.error(f"Error saving data to Excel: {e}")
        raise  # Forward the exception to fetch_data function, No need to return false.
        # return False


def save_to_excel_with_pandas(indices_data, excel_path):
    try:
        indices_data.to_excel(excel_path, sheet_name='AllIndicesDataPandas', index=False)
        logging.info(f"Data successfully saved to {excel_path}")
    except Exception as e:
        logging.info(f"Error saving data to Excel: {e}")
        raise # Forward the exception to fetch_data function


def fetch_and_write_data(excel_file_path):
    start_time = time.time()
    try:
        logging.info("Fetching market index data")
        indices_data = all_indices_data()
        logging.info("Data fetched successfully")
        logging.info(f"Fetching Indices data from NSE Time : {time.time() - start_time:.4f} seconds")

        write_indices_to_excel(indices_data, excel_file_path)
        # save_to_excel_with_pandas(indices_data, excel_file_path)

        logging.info("Data successfully saved to Excel.")

        return True, "Data Fetched Successfully" , "#04941c"

    except Exception as e:
        logging.error(f'Error occurred: {e}')
        logging.error(f"Error while Fetching Indices data from NSE, Time taken : {time.time() - start_time:.4f} seconds")
        return False , "Error while fetching data. Please try again" , "#fa1e43"



def write_indices_to_excel_old(indices_data, excel_path):

    try:
        logging.info("Saving data to Excel Sheet function triggered")

        excel_workbook = load_workbook(excel_path)

        if "AllIndicesData" in excel_workbook.sheetnames:
            sheet = excel_workbook["AllIndicesData"]
        else:
            sheet = excel_workbook.create_sheet("AllIndicesData")
            for col_idx, column in enumerate(indices_data.columns, start=1):
                sheet.cell(row=1, column=col_idx, value=column)

        for row_idx, row in enumerate(indices_data.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        excel_workbook.save(excel_path)
        logging.info(f"Data successfully saved to: {excel_path}")

    except Exception as e:
        logging.error(f"Error saving data to Excel: {e}")



# JSON Handling code


# def save_temp_file_path(filename):
#     config_file_path = os.path.join(get_base_path(), "config","settings.json")
#     with open(config_file_path, 'r+') as f:
#         data = json.load(f)
#
#         if not data['previous_temp_file_paths']:
#             data['previous_temp_file_paths'] = {}
#             max_key = 0
#         else:
#             max_key = max(int(key) for key in data['previous_temp_file_paths'].keys()) + 1
#
#         data['previous_temp_file_paths'][str(max_key)] = filename
#
#         f.seek(0)
#         json.dump(data, f, indent=4)
#         f.truncate()


# Playwright Code

# def find_chromium_folder_name(playwright_directory_path):
#     for root, dirs, files in os.walk(playwright_directory_path):
#         for dir_name in dirs:
#             if dir_name.startswith('chromium_headless_shell'):
#                 return os.path.join(root, dir_name)
#     return ""
#
#
# def get_browser_binaries_path():
#     binaries_path = ""
#     try:
#         if getattr(sys, "_MEIPASS", False):
#             logging.info('Getting path for browser binaries')
#             app_base_path = os.path.abspath(getattr(sys, "_MEIPASS", ""))
#             temp_dir_path = os.path.dirname(app_base_path)
#             local_dir_path = os.path.dirname(temp_dir_path)
#             playwright_dir_path = os.path.join(local_dir_path, "ms-playwright")
#
#             chrome_binaries_dir_name = find_chromium_folder_name(playwright_dir_path)
#
#             if chrome_binaries_dir_name:
#                 binaries_path = os.path.join(playwright_dir_path, chrome_binaries_dir_name, "chrome-win", "headless_shell.exe")
#                 logging.info(f"Browser binaries path determined: {binaries_path}")
#             else:
#                 logging.warning("Chromium folder not found in ms-playwright directory")
#
#     except Exception as e:
#         logging.error(f"Error while getting browser binaries path: {e}")
#
#     return binaries_path
#
# browser_binaries_path = get_browser_binaries_path()
# logging.info(f"Final browser binaries path: {browser_binaries_path}")
#
#
#
# def fetch_data_by_playwright(excel_file_path):
#     try:
#         save_directory = os.path.join(get_base_path(), "assets", "csv")
#         os.makedirs(save_directory, exist_ok=True)
#         save_file_path = os.path.join(save_directory, "all_indices.csv")
#
#         with sync_playwright() as p:
#             logging.info('Scrapping function triggered')
#             if getattr(sys, "_MEIPASS", False):
#                 logging.info("If ex")
#                 browser = p.chromium.launch(
#                     executable_path=browser_binaries_path,
#                     headless=True,
#                     args=["--disable-http2"])
#             else:
#                 browser = p.chromium.launch(headless=True,args=["--disable-http2"])
#             context = browser.new_context(
#                 user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
#                 bypass_csp=True,
#                 ignore_https_errors=True
#             )
#             page = context.new_page()
#             logging.info('Browser launched')
#             page.goto('https://www.google.co.in')
#             logging.info('Visited url')
#
#             h1_text = page.locator("a.MV3Tnb").first.text_content()
#
#             workbook = load_workbook(excel_file_path)
#             logging.info('Excel sheet loaded')
#             sheet = workbook.active
#             logging.info('Opened sheet')
#
#             sheet["A1"] = h1_text
#
#             workbook.save(excel_file_path)
#             logging.info('Success')
#
#             browser.close()
#
#     except Exception as e:
#         logging.info(f'Error occured : {e}')
#         print(f"An error occurred: {e}")


# Mapping function code


# def process_excel_data(temp_workbook, excel_path):
#
#     try:
#         logging.info("Process Excel Data function triggered")
#         if os.path.exists(excel_path):
#             excel_workbook = load_workbook(excel_path)
#         else:
#             excel_workbook = Workbook()
#
#         excel_sheet = None
#         for existing_sheet in excel_workbook.worksheets:
#             if existing_sheet.title == "AllIndicesDataNew":
#                 excel_sheet = existing_sheet
#                 break
#
#         if excel_sheet is None:
#             excel_sheet = excel_workbook.create_sheet("AllIndicesDataNew")
#
#         temp_sheet = temp_workbook.active
#         # Column mapping
#         column_mapping = {
#             'B': 'A', 'C': 'B', 'D': 'C',
#             'F': 'D', 'G': 'E', 'H': 'F',
#             'I': 'G', 'J': 'H', 'U': 'I'
#         }
#
#         # Copy data based on column mapping
#         for row_idx in range(1, temp_sheet.max_row + 1):
#             for temp_col, dest_col in column_mapping.items():
#                 temp_value = temp_sheet[f"{temp_col}{row_idx}"].value  # Read from temp column
#                 excel_sheet[f"{dest_col}{row_idx}"].value = temp_value  # Write to destination column
#
#         # Save the updated destination workbook
#         os.makedirs(os.path.dirname(excel_path), exist_ok=True)
#         excel_workbook.save(excel_path)
#         logging.info(f"Data processed successfully into: {excel_path}")
#
#     except Exception as e:
#         logging.error(f"Error processing Excel data: {e}")
