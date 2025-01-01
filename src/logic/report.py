import logging
import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime
from src.logic.common_utils import get_report_date


def generate_daily_report(excel_file_path, report_dir):
    try:
        title_column = "A"
        buy_column = "PQ"
        sell_column = "RD"
        buy_target_columns = ["PT", "PV", "PX", "PZ", "QA", "QE", "QF", "QG", "QH"]
        sell_target_columns = ["RG", "RI", "RL", "RM", "RN", "RO", "SA", "SB"]
        positional_sl_column = "QZ"

        # Row numbers
        row_numbers = [2, 21, 23, 17, 3, 6, 13, 22, 25, 26, 28, 30, 31, 32, 33, 34, 35, 76, 77, 78, 80, 83, 84]

        wb = load_workbook(excel_file_path, data_only=True)
        if "ALL NEW (CLOSE)" not in wb.sheetnames:
            return False, "All New (CLOSE) Sheet was not found", "#fa1e43"
        sheet = wb["ALL NEW (CLOSE)"]

        os.makedirs(report_dir, exist_ok=True)

        current_time = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
        # report_title_time = datetime.now().strftime("%d %b %Y")
        report_file_name = f"Daily_Reports_{current_time}.txt"
        report_file_path = os.path.join(report_dir, report_file_name)

        # Helper function to format cell value
        def format_cell_value(cell_value):
            if isinstance(cell_value, (int, float)):
                return int(cell_value)
            return "-" if cell_value is None else cell_value

        report_lines = [f"Daily Reports {get_report_date()}\n\n\n"]  # Report data Array initialized with title

        for row in row_numbers:
            # Extract values for the current row
            title = sheet[f"{title_column}{row}"].value
            title = format_cell_value(title)
            buy = format_cell_value(sheet[f"{buy_column}{row}"].value)
            sell = format_cell_value(sheet[f"{sell_column}{row}"].value)
            buy_targets = [
                str(format_cell_value(sheet[f"{col}{row}"].value))
                for col in buy_target_columns
            ]
            sell_targets = [
                str(format_cell_value(sheet[f"{col}{row}"].value))
                for col in sell_target_columns
            ]
            positional_sl = format_cell_value(sheet[f"{positional_sl_column}{row}"].value)

            # Format the record
            # \n for new line \t for tab
            record = (
                f"{title}-\n"
                f"Buy:-{buy} Target {','.join(buy_targets)}\n"
                f"Sell:-{sell} Target {','.join(sell_targets)}\n"
                f"{title}  Positional sl  {positional_sl}\n\n"
            )
            report_lines.append(record)

        with open(report_file_path, "w") as report_file:
            report_file.writelines(report_lines)

        return True, "Daily Report generated successfully.", "#04941c"  # Success

    except Exception as e:
        logging.error(f"Error: {e}")
        return False, "Error occurred during Generate daily report function.", "#fa1e43"


def generate_weekly_report(excel_file_path, report_dir):
    try:
        title_column = "A"
        buy_column = "PQ"
        sell_column = "QX"
        buy_target_columns = ["PT", "PX", "PZ", "QA", "QC", "QB"]
        sell_target_columns = ["RA", "RE", "RG", "RH", "RJ"]

        # Row numbers
        row_numbers_mapping = {2: "NIFTY 50", 20: "BANK NIFTY", 22: "FIN NIFTY", 17: "MIDCAP NIFTY", 3: "NIFTY NEXT"}
        row_numbers = [2, 21, 23, 17, 3]

        wb = load_workbook(excel_file_path, data_only=True)
        if "WEEKLY" not in wb.sheetnames:
            return False, "WEEKLY Sheet was not found", "#fa1e43"
        sheet = wb["WEEKLY"]

        os.makedirs(report_dir, exist_ok=True)

        current_time = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
        # report_title_time = datetime.now().strftime("%d %b %Y")
        report_file_name = f"Weekly_Reports_{current_time}.txt"
        report_file_path = os.path.join(report_dir, report_file_name)

        # Helper function to format cell Value
        def format_cell_value(cell_value):
            if isinstance(cell_value, (int, float)):
                return int(cell_value)
            return "-" if cell_value is None else cell_value

        report_lines = [f"Weekly Reports {get_report_date()}\n\n\n"]  # Report data Array initialized with title

        for row in row_numbers:
            title = f"{title_column}{row} Weekly Levels"
            buy = format_cell_value(sheet[f"{buy_column}{row}"].value)
            sell = format_cell_value(sheet[f"{sell_column}{row}"].value)
            buy_targets = [
                str(format_cell_value(sheet[f"{col}{row}"].value))
                for col in buy_target_columns
            ]
            sell_targets = [
                str(format_cell_value(sheet[f"{col}{row}"].value))
                for col in sell_target_columns
            ]

            # Format the record
            record = (
                f"{title}-\n"
                f"Buy:-{buy} Target {','.join(buy_targets)}\n"
                f"Sell:-{sell} Target {','.join(sell_targets)}\n\n"
            )

            report_lines.append(record)

        with open(report_file_path, "w") as report_file:
            report_file.writelines(report_lines)

        return True, "Weekly Report generated successfully.", "#04941c"  # Success

    except Exception as e:
        logging.error(f"Error: {e}")
        return False, "Error occurred during Generate weekly report function.", "#fa1e43"


def generate_monthly_report(excel_file_path, report_dir):
    try:
        title_column = "RT"
        buy_column = "PQ"
        sell_column = "QX"
        buy_target_columns = ["PR", "PT", "PV", "PX", "PY", "PZ", "QA", "QB", "QC"]
        sell_target_columns = ["QY", "RA", "RC", "RE", "RF", "RG", "RI", "RH", "RJ"]

        # Row numbers
        row_numbers_mapping = {2: "NIFTY 50", 21: "BANK NIFTY", 23: "FIN NIFTY", 17: "MIDCAP NIFTY", 3: "NIFTY NEXT"}

        wb = load_workbook(excel_file_path, data_only=True)
        if "MONTHLY" not in wb.sheetnames:
            return False, "MONTHLY Sheet was not found", "#fa1e43"
        sheet = wb["MONTHLY"]

        os.makedirs(report_dir, exist_ok=True)

        current_time = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
        report_title_time = datetime.now().strftime("%d %b %Y")
        report_file_name = f"Monthly_Reports_{current_time}.txt"
        report_file_path = os.path.join(report_dir, report_file_name)

        # Helper function to format cell value
        def format_cell_value(cell_value):
            if isinstance(cell_value, (int, float)):
                return int(cell_value)
            return "-" if cell_value is None else cell_value

        report_lines = [f"Monthly Reports {report_title_time}\n\n\n"]  # Report data Array initialized with title
        for row, row_title in row_numbers_mapping.items():
            title = f"{format_cell_value(sheet[f"{title_column}{row}"].value)} MONTHLY LEVELS"
            # title = f"{row_title} MONTHLY LEVELS"
            buy = format_cell_value(sheet[f"{buy_column}{row}"].value)
            sell = format_cell_value(sheet[f"{sell_column}{row}"].value)
            buy_targets = [
                str(format_cell_value(sheet[f"{col}{row}"].value))
                for col in buy_target_columns
            ]
            sell_targets = [
                str(format_cell_value(sheet[f"{col}{row}"].value))
                for col in sell_target_columns
            ]

            record = (
                f"{title}-\n"
                f"Buy:-{buy} Target {','.join(buy_targets)}\n"
                f"Sell:-{sell} Target {','.join(sell_targets)}\n\n"
            )
            report_lines.append(record)

        with open(report_file_path, "w") as report_file:
            report_file.writelines(report_lines)

        return True, "Monthly Report generated successfully.", "#04941c"  # Success

    except Exception as e:
        logging.error(f"Error: {e}")
        return False, "Error occurred during Generate monthly report function.", "#fa1e43"


def generate_report_pandas(excel_file_path, report_dir):
    try:
        title_column = "A"
        buy_column = "PQ"
        sell_column = "RD"
        buy_target_columns = ["PT", "PV", "PX", "PZ", "QA", "QD", "QE", "QF"]
        sell_target_columns = ["RG", "RI", "RK", "RM", "RN", "RO", "SA", "SB"]
        positional_sl_column = "QZ"

        # Row numbers
        row_numbers = [2, 20, 22, 17, 3, 21, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34,
                       55, 56, 62, 63, 6, 13, 14]

        try:
            df = pd.read_excel(excel_file_path, sheet_name="ALL NEW (CLOSE)", engine="openpyxl")
        except ValueError:
            return False

        os.makedirs(report_dir, exist_ok=True)

        current_time = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
        report_title_time = datetime.now().strftime("%d %b %Y")
        report_file_name = f"Reports_{current_time}.txt"
        report_file_path = os.path.join(report_dir, report_file_name)

        def get_integer_value(cell_value):
            if pd.isna(cell_value):
                return "-"
            if isinstance(cell_value, (int, float)):
                return int(cell_value)
            return cell_value

        report_lines = [f"Reports {report_title_time}\n\n"]
        for row in row_numbers:
            row_data = df.iloc[row - 1]  # Pandas is 0-indexed; Excel is 1-indexed
            title = get_integer_value(row_data[title_column])
            buy = get_integer_value(row_data[buy_column])
            sell = get_integer_value(row_data[sell_column])
            buy_targets = [str(get_integer_value(row_data[col])) for col in buy_target_columns]
            sell_targets = [str(get_integer_value(row_data[col])) for col in sell_target_columns]
            positional_sl = get_integer_value(row_data[positional_sl_column])

            record = (
                f"{title}-\n"
                f"Buy:-{buy} Target {','.join(buy_targets)}\n"
                f"Sell:-{sell} Target {','.join(sell_targets)}\n"
                f"{title} Positional sl\t\t{positional_sl}\n\n"
            )
            report_lines.append(record)

        with open(report_file_path, "w") as report_file:
            report_file.writelines(report_lines)

        return True  # Success

    except Exception as e:
        print(f"Error: {e}")
        return False
